from flask import Flask, request, jsonify, render_template, session, redirect, url_for, make_response, send_file
from authlib.integrations.flask_client import OAuth
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3
import secrets
import time
import os
import re

import llm  # optional Gemini helpers (reads its key lazily, so import order is fine)
import embeddings  # semantic-similarity foundation (also degrades to no-op without a key)

# Load GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET / SECRET_KEY from a local .env file
# (if present) so they persist across restarts without re-exporting them each time.
# Real environment variables still take precedence over .env values.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
# Signs the session cookie that keeps a user logged in. Set a real value in production.
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,   # JS can't read the cookie
    SESSION_COOKIE_SAMESITE="Lax",  # not sent on cross-site POSTs (CSRF defence in depth)
    # Send the cookie only over HTTPS in production; off by default for local http dev.
    SESSION_COOKIE_SECURE=os.environ.get("COOKIE_SECURE", "").lower() in ("1", "true", "yes"),
    # Don't cache static files in dev, so CSS/JS edits show up on a normal refresh.
    SEND_FILE_MAX_AGE_DEFAULT=0,
)
DB = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "tips.db"))

# ── Google OAuth — only enabled when credentials are present, so the app still
# runs (just without login) until you set GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET. ──
oauth = OAuth(app)
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")
AUTH_ENABLED = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)
if AUTH_ENABLED:
    oauth.register(
        name="google",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={"scope": "openid email profile"},
    )


def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def current_user_id():
    """The logged-in user's id, or None. Reads the signed session cookie."""
    return session.get("uid")


# ── Administrator login (separate from Google user login) ──
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
# Store/compare a HASH, never plaintext. Prefer a pre-computed ADMIN_PASSWORD_HASH;
# otherwise hash ADMIN_PASSWORD (default "admin" for local dev) once at startup.
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH") or generate_password_hash(
    os.environ.get("ADMIN_PASSWORD", "admin"), method="pbkdf2:sha256"
)

# Simple in-memory rate limit for admin login (per process; enough for dev).
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW = 300  # seconds
_login_attempts = {}  # ip -> [timestamps of recent failures]


def _login_blocked(ip):
    now = time.time()
    recent = [t for t in _login_attempts.get(ip, []) if now - t < LOGIN_WINDOW]
    _login_attempts[ip] = recent
    return len(recent) >= LOGIN_MAX_ATTEMPTS


def _login_failed(ip):
    _login_attempts.setdefault(ip, []).append(time.time())


def is_admin():
    return bool(session.get("is_admin"))


def admin_required(fn):
    """Block the endpoint unless the session is an administrator."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_admin():
            return jsonify({"error": "Administrator access required."}), 403
        return fn(*args, **kwargs)
    return wrapper


# ── CSRF protection ──
# A token is kept in the (signed) session and echoed to the page via /api/me. The
# front-end sends it back as an X-CSRF-Token header on every state-changing request.
# A cross-site page can't read the token, so it can't forge those requests.
def csrf_token():
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_hex(32)
        session["csrf_token"] = token
    return token


@app.before_request
def csrf_protect():
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        expected = session.get("csrf_token", "")
        sent = request.headers.get("X-CSRF-Token", "")
        if not expected or not sent or not secrets.compare_digest(sent, expected):
            return jsonify({"error": "Missing or invalid CSRF token — reload the page."}), 400


MIGRATIONS_DIR = os.path.join(os.path.dirname(__file__), "migrations")


def run_migrations():
    """Apply any migration .sql files (migrations/NNN_*.sql) not yet recorded, in order.

    Each applied file is recorded in schema_migrations so it runs exactly once. Files use
    `CREATE TABLE IF NOT EXISTS`, so this is also safe to run against an existing database.
    """
    with get_db() as conn:
        conn.execute(
            "CREATE TABLE IF NOT EXISTS schema_migrations ("
            "  filename TEXT PRIMARY KEY, applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"
        )
        applied = {r["filename"] for r in conn.execute("SELECT filename FROM schema_migrations")}
        files = sorted(f for f in os.listdir(MIGRATIONS_DIR) if f.endswith(".sql"))
        for fname in files:
            if fname in applied:
                continue
            with open(os.path.join(MIGRATIONS_DIR, fname)) as fh:
                conn.executescript(fh.read())
            conn.execute("INSERT INTO schema_migrations (filename) VALUES (?)", (fname,))
            app.logger.info("applied migration %s", fname)
        conn.commit()


def init_db():
    run_migrations()


def get_or_create_tag(conn, name, tier=None):
    """Return the tag id, creating it if needed.

    New tags default to 'primary'. If a tier is explicitly given, it is applied
    (also reclassifying an existing tag).
    """
    name = name.strip().lower()
    row = conn.execute("SELECT id FROM tags WHERE name = ?", (name,)).fetchone()
    if row:
        if tier in ("primary", "secondary"):
            conn.execute("UPDATE tags SET tier = ? WHERE id = ?", (tier, row["id"]))
        return row["id"]
    cur = conn.execute(
        "INSERT INTO tags (name, tier) VALUES (?, ?)", (name, tier or "primary")
    )
    return cur.lastrowid


def will_have_primary(conn, tag_names):
    """True if attaching these tags would give the tip at least one primary tag.

    A name that doesn't exist yet counts as primary (new tags default to primary).
    Safe to call before any writes — it does not mutate.
    """
    for name in tag_names:
        name = name.strip().lower()
        if not name:
            continue
        row = conn.execute("SELECT tier FROM tags WHERE name = ?", (name,)).fetchone()
        if row is None or row["tier"] == "primary":
            return True
    return False


def video_embed(url, start=0, end=0):
    """Turn a YouTube / Vimeo / Cloudflare Stream URL into an embeddable player src.

    Returns the iframe src, or None if the URL isn't a recognised video link. We only ever build
    the src from an id captured out of a known host, so the result is safe to drop into an iframe.

    Optional start/end are whole seconds (0 = unset). YouTube honours both; Vimeo and Cloudflare
    Stream honour the start only (a hard stop on those would need their JS player SDK).
    """
    url = (url or "").strip()
    if not url:
        return None
    try:
        start = max(0, int(start or 0))
        end = max(0, int(end or 0))
    except (TypeError, ValueError):
        start = end = 0
    m = re.search(r"(?:youtube\.com/(?:watch\?v=|embed/|shorts/|v/)|youtu\.be/)([A-Za-z0-9_-]{6,})", url)
    if m:
        src = "https://www.youtube-nocookie.com/embed/%s?rel=0" % m.group(1)
        if start:
            src += "&start=%d" % start
        if end and end > start:
            src += "&end=%d" % end
        return src
    m = re.search(r"vimeo\.com/(?:video/)?(\d+)", url)
    if m:
        src = "https://player.vimeo.com/video/%s" % m.group(1)
        return src + ("#t=%ds" % start if start else "")
    m = re.search(r"cloudflarestream\.com/([A-Za-z0-9]+)", url)
    if m:
        src = "https://iframe.cloudflarestream.com/%s" % m.group(1)
        return src + ("?startTime=%ds" % start if start else "")
    return None


def tip_with_tags(conn, tip_id):
    tip = conn.execute("SELECT * FROM tips WHERE id = ?", (tip_id,)).fetchone()
    if not tip:
        return None
    tags = conn.execute(
        "SELECT t.name FROM tags t JOIN tip_tags tt ON t.id = tt.tag_id WHERE tt.tip_id = ?",
        (tip_id,),
    ).fetchall()
    # Vote tally for everyone; the current user's own vote + favorite status if logged in.
    score = conn.execute(
        "SELECT COALESCE(SUM(value), 0) AS s FROM votes WHERE tip_id = ?", (tip_id,)
    ).fetchone()["s"]
    my_vote = 0
    uid = current_user_id()
    if uid:
        v = conn.execute(
            "SELECT value FROM votes WHERE tip_id = ? AND user_id = ?", (tip_id, uid)
        ).fetchone()
        my_vote = v["value"] if v else 0
    favorited = my_vote == 1  # a tip is "favorited" exactly when the user has upvoted it
    video_url = tip["video_url"] or ""
    video_start = tip["video_start"] or 0
    video_end = tip["video_end"] or 0
    # Admin-written analysis overrides, keyed by lens (only the non-empty ones).
    analysis = {
        r["lens"]: r["text"]
        for r in conn.execute(
            "SELECT lens, text FROM tip_analysis WHERE tip_id = ? AND text != ''", (tip_id,)
        ).fetchall()
    }
    return {
        "id": tip["id"],
        "content": tip["content"],
        "anecdote": tip["anecdote"] or "",
        "tags": [r["name"] for r in tags],
        "score": score,
        "my_vote": my_vote,
        "favorited": favorited,
        "video_url": video_url,
        "video_start": video_start,
        "video_end": video_end,
        "video_embed": video_embed(video_url, video_start, video_end),
        "analysis": analysis,
    }


def embed_quietly(conn, tip_id, content, anecdote=""):
    """Refresh a tip's embedding, but never let an embedding/network error break the write.

    Content edits make a tip's stored vector stale; this keeps the index current. If the API
    is down or unconfigured the tip is simply left for the next 'rebuild embeddings' run.
    """
    if not embeddings.is_enabled():
        return
    try:
        embeddings.store_one(conn, tip_id, content, anecdote)
    except llm.LLMError as e:
        app.logger.warning("embedding skipped for tip %s: %s", tip_id, e)


@app.get("/")
def index():
    # Never cache the page itself, so edits show up on a normal refresh.
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/api/tips")
def search_tips():
    tags_param = request.args.get("tags", "").strip()
    favorites_only = request.args.get("favorites") == "1"
    uid = current_user_id()
    with get_db() as conn:
        if tags_param:
            tag_list = [t.strip().lower() for t in tags_param.split(",") if t.strip()]
            placeholders = ",".join("?" * len(tag_list))
            # tips that have ALL the requested tags
            rows = conn.execute(
                f"""
                SELECT t.id FROM tips t
                JOIN tip_tags tt ON t.id = tt.tip_id
                JOIN tags tg ON tt.tag_id = tg.id
                WHERE tg.name IN ({placeholders})
                GROUP BY t.id
                HAVING COUNT(DISTINCT tg.name) = ?
                ORDER BY t.created_at DESC
                """,
                (*tag_list, len(tag_list)),
            ).fetchall()
        else:
            rows = conn.execute("SELECT id FROM tips ORDER BY created_at DESC").fetchall()

        ids = [r["id"] for r in rows]
        if favorites_only:
            if not uid:
                return jsonify([])  # not signed in → no favorites
            fav = {r["tip_id"] for r in conn.execute(
                "SELECT tip_id FROM votes WHERE user_id = ? AND value = 1", (uid,)
            ).fetchall()}
            ids = [i for i in ids if i in fav]

        return jsonify([tip_with_tags(conn, i) for i in ids])


@app.get("/api/tips/search")
def semantic_search():
    """Rank tips by semantic similarity to a free-text query.

    Open to everyone (it's a discovery feature). Returns {"enabled": bool, "results": [...]}
    where each result is a full tip plus a "similarity" score in 0..1. When embeddings aren't
    configured, enabled is False and results is empty (the UI then hides the 'Meaning' option).
    """
    q = (request.args.get("q") or "").strip()
    try:
        k = max(1, min(int(request.args.get("k", 20)), 50))
    except ValueError:
        k = 20
    if not q or not embeddings.is_enabled():
        return jsonify({"enabled": embeddings.is_enabled(), "results": []})
    with get_db() as conn:
        try:
            hits = embeddings.search(conn, q, k=k)
        except llm.LLMError as e:
            return jsonify({"error": "Search failed: %s" % e}), 502
        results = []
        for h in hits:
            tip = tip_with_tags(conn, h["tip_id"])
            if tip:
                tip["similarity"] = h["score"]
                results.append(tip)
    return jsonify({"enabled": True, "results": results})


def _fts_match(q):
    """Turn free user text into a safe FTS5 MATCH expression.

    We only keep word characters (dropping FTS operators/punctuation that could be a syntax
    error), quote each token, and add a prefix `*` so partial words match. Space-separated
    tokens are AND-ed by FTS5, so all words must appear. Returns None for an empty query.
    """
    tokens = re.findall(r"\w+", q.lower())
    if not tokens:
        return None
    return " ".join('"%s"*' % t for t in tokens)


@app.get("/api/tips/fts")
def fulltext_search():
    """Search the actual words of tips (content + anecdote) via SQLite FTS5, ranked by relevance.

    Open to everyone, needs no API key. Returns {"results": [...full tips...]} best-match first.
    """
    q = (request.args.get("q") or "").strip()
    try:
        k = max(1, min(int(request.args.get("k", 50)), 100))
    except ValueError:
        k = 50
    match = _fts_match(q)
    if not match:
        return jsonify({"results": []})
    with get_db() as conn:
        try:
            rows = conn.execute(
                "SELECT rowid FROM tips_fts WHERE tips_fts MATCH ? ORDER BY rank LIMIT ?",
                (match, k),
            ).fetchall()
        except sqlite3.OperationalError as e:
            return jsonify({"error": "Search failed: %s" % e}), 500
        results = [t for t in (tip_with_tags(conn, r["rowid"]) for r in rows) if t]
    return jsonify({"results": results})


@app.post("/api/advise")
def advise():
    """Retrieve the most relevant tips for a described situation, then synthesise grounded
    advice that cites them (RAG). Open to everyone. Returns {answer, used:[ids], tips:[...]}.
    """
    if not embeddings.is_enabled():
        return jsonify({"error": "The advice assistant needs an AI key (set GEMINI_API_KEY)."}), 503
    situation = ((request.get_json(force=True) or {}).get("situation") or "").strip()
    if not situation:
        return jsonify({"error": "Describe your situation first."}), 400
    with get_db() as conn:
        try:
            hits = embeddings.search(conn, situation, k=6)
        except llm.LLMError as e:
            return jsonify({"error": "Retrieval failed: %s" % e}), 502
        tips = []
        for h in hits:
            tip = tip_with_tags(conn, h["tip_id"])
            if tip:
                tip["similarity"] = h["score"]
                tips.append(tip)
    if not tips:
        return jsonify({"answer": "There aren't any tips to draw on yet.", "used": [], "tips": []})
    try:
        result = llm.advise(situation, [{"id": t["id"], "content": t["content"]} for t in tips])
    except llm.LLMError as e:
        return jsonify({"error": "Advice generation failed: %s" % e}), 502
    return jsonify({"answer": result["answer"], "used": result["used"], "tips": tips})


@app.get("/api/tips/<int:tip_id>/related")
def related_tips(tip_id):
    """The tips most similar in meaning to this one (uses the stored vector — no API call).

    Powers the semantic 'next suggested tip' and the network's related-link mode. Returns
    {"enabled": bool, "related": [{"tip_id", "score"}...]} ordered most-similar first.
    """
    try:
        k = max(1, min(int(request.args.get("k", 12)), 50))
    except ValueError:
        k = 12
    if not embeddings.is_enabled():
        return jsonify({"enabled": False, "related": []})
    with get_db() as conn:
        if not conn.execute("SELECT 1 FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        return jsonify({"enabled": True, "related": embeddings.neighbors(conn, tip_id, k=k)})


@app.post("/api/tips/<int:tip_id>/analyze")
def analyze_tip_lens(tip_id):
    """Analyse one tip through a chosen lens (how to apply, when not to, opposing wisdom, common
    misreadings, notable figures). On-demand — the caller picks the lens.

    If an admin has written text for this lens it's returned as {"text": ..., "custom": true}
    (no LLM call). Otherwise the lens is generated on the fly as {"points": [...]}.
    """
    lens = ((request.get_json(force=True) or {}).get("lens") or "").strip()
    if lens not in llm.ANALYSIS_LENSES:
        return jsonify({"error": "Unknown analysis option."}), 400
    with get_db() as conn:
        row = conn.execute("SELECT content FROM tips WHERE id = ?", (tip_id,)).fetchone()
        if not row:
            return jsonify({"error": "tip not found"}), 404
        override = conn.execute(
            "SELECT text FROM tip_analysis WHERE tip_id = ? AND lens = ?", (tip_id, lens)
        ).fetchone()
    if override and override["text"].strip():
        return jsonify({"text": override["text"], "custom": True})
    if not llm.is_enabled():
        return jsonify({"error": "AI analysis isn't configured (set GROQ_API_KEY)."}), 503
    try:
        return jsonify(llm.analyze_tip(row["content"], lens))
    except llm.LLMError as e:
        return jsonify({"error": "Analysis failed: %s" % e}), 502


@app.put("/api/tips/<int:tip_id>/analysis")
@admin_required
def update_tip_analysis(tip_id):
    """Set (or clear) the admin's own "choose an angle" text for a tip. Body:
    {"analysis": {"apply": "...", "avoid": "", ...}}. Empty / missing lenses are cleared so the
    reader falls back to on-demand AI generation for those angles. Admin only."""
    overrides = (request.get_json(force=True) or {}).get("analysis") or {}
    with get_db() as conn:
        if not conn.execute("SELECT id FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        for lens in llm.ANALYSIS_LENSES:
            text = (overrides.get(lens) or "").strip()
            if text:
                conn.execute(
                    "INSERT INTO tip_analysis (tip_id, lens, text) VALUES (?, ?, ?) "
                    "ON CONFLICT(tip_id, lens) DO UPDATE SET text = excluded.text",
                    (tip_id, lens, text),
                )
            else:
                conn.execute(
                    "DELETE FROM tip_analysis WHERE tip_id = ? AND lens = ?", (tip_id, lens)
                )
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


@app.post("/api/tips")
@admin_required
def create_tip():
    data = request.get_json(force=True)
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "content is required"}), 400
    anecdote = (data.get("anecdote") or "").strip()
    tag_names = [t for t in data.get("tags", []) if t.strip()]

    with get_db() as conn:
        if not will_have_primary(conn, tag_names):
            return jsonify({"error": "Each tip needs at least one primary tag."}), 400
        cur = conn.execute(
            "INSERT INTO tips (content, anecdote) VALUES (?, ?)", (content, anecdote)
        )
        tip_id = cur.lastrowid
        for name in tag_names:
            tag_id = get_or_create_tag(conn, name)
            conn.execute(
                "INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)", (tip_id, tag_id)
            )
        embed_quietly(conn, tip_id, content, anecdote)
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id)), 201


@app.put("/api/tips/<int:tip_id>")
@admin_required
def update_tip(tip_id):
    data = request.get_json(force=True)
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "content is required"}), 400
    anecdote = (data.get("anecdote") or "").strip()

    with get_db() as conn:
        if not conn.execute("SELECT id FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        conn.execute(
            "UPDATE tips SET content = ?, anecdote = ? WHERE id = ?",
            (content, anecdote, tip_id),
        )
        embed_quietly(conn, tip_id, content, anecdote)  # text changed → refresh its vector
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


@app.post("/api/tips/<int:tip_id>/video")
@admin_required
def set_tip_video(tip_id):
    """Attach (or clear) a YouTube / Vimeo / Cloudflare Stream video on a tip, with optional
    start/stop times (seconds). Admin only."""
    data = request.get_json(force=True) or {}
    url = (data.get("video_url") or "").strip()
    if url and not video_embed(url):
        return jsonify({"error": "Unrecognised video link — use a YouTube, Vimeo, or Cloudflare Stream URL."}), 400

    def _secs(v):
        try:
            return max(0, int(v or 0))
        except (TypeError, ValueError):
            return 0
    start, end = (_secs(data.get("video_start")), _secs(data.get("video_end"))) if url else (0, 0)

    with get_db() as conn:
        if not conn.execute("SELECT id FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        conn.execute("UPDATE tips SET video_url = ?, video_start = ?, video_end = ? WHERE id = ?",
                     (url, start, end, tip_id))
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


@app.delete("/api/tips/<int:tip_id>")
@admin_required
def delete_tip(tip_id):
    with get_db() as conn:
        if not conn.execute("SELECT id FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        conn.execute("DELETE FROM tips WHERE id = ?", (tip_id,))
        conn.commit()
        return jsonify({"deleted": tip_id})


@app.delete("/api/tips")
@admin_required
def clear_all_tips():
    """Delete every tip. Cascades to its tags-links, votes, favorites, seen marks, embeddings
    and angle overrides, and the FTS index is cleared by the AFTER DELETE trigger. The tag
    vocabulary itself is kept. Irreversible — the UI double-confirms before calling this."""
    with get_db() as conn:
        n = conn.execute("SELECT COUNT(*) AS c FROM tips").fetchone()["c"]
        conn.execute("DELETE FROM tips")
        conn.commit()
    return jsonify({"deleted": n})


@app.delete("/api/tags")
@admin_required
def clear_all_tags():
    """Delete the entire tag vocabulary. Cascades to remove every tip↔tag link, so the tips
    themselves remain but lose their tags. Irreversible — the UI double-confirms."""
    with get_db() as conn:
        n = conn.execute("SELECT COUNT(*) AS c FROM tags").fetchone()["c"]
        conn.execute("DELETE FROM tags")
        conn.commit()
    return jsonify({"deleted": n})


@app.put("/api/tips/<int:tip_id>/tags")
@admin_required
def update_tags(tip_id):
    data = request.get_json(force=True)
    tag_names = [t.strip().lower() for t in data.get("tags", []) if t.strip()]

    with get_db() as conn:
        if not conn.execute("SELECT id FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        if not will_have_primary(conn, tag_names):
            return jsonify({"error": "Each tip needs at least one primary tag."}), 400
        conn.execute("DELETE FROM tip_tags WHERE tip_id = ?", (tip_id,))
        for name in tag_names:
            tag_id = get_or_create_tag(conn, name)
            conn.execute(
                "INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)", (tip_id, tag_id)
            )
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


@app.post("/api/tips/<int:tip_id>/tags")
@admin_required
def add_tag_to_tip(tip_id):
    """Add a single tag to one tip (creating it if new). For the 'apply a tag' tool."""
    data = request.get_json(force=True) or {}
    name = (data.get("tag") or "").strip().lower()
    tier = data.get("tier", "secondary")
    if tier not in ("primary", "secondary"):
        tier = "secondary"
    if not name:
        return jsonify({"error": "tag is required"}), 400
    with get_db() as conn:
        if not conn.execute("SELECT 1 FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        # Reuse an existing tag as-is (don't change its tier); only new tags use `tier`.
        existing = conn.execute("SELECT id FROM tags WHERE name = ?", (name,)).fetchone()
        tag_id = existing["id"] if existing else get_or_create_tag(conn, name, tier=tier)
        conn.execute("INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)", (tip_id, tag_id))
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


@app.delete("/api/tips/<int:tip_id>/tags/<name>")
@admin_required
def remove_tag_from_tip(tip_id, name):
    """Remove a single tag from one tip, unless it would leave the tip with no primary tag."""
    name = name.strip().lower()
    with get_db() as conn:
        if not conn.execute("SELECT 1 FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        remaining = [r["name"] for r in conn.execute(
            "SELECT t.name FROM tags t JOIN tip_tags tt ON t.id = tt.tag_id "
            "WHERE tt.tip_id = ? AND t.name <> ?", (tip_id, name)
        )]
        if not will_have_primary(conn, remaining):
            return jsonify({"error": "That would leave the tip with no primary tag."}), 400
        tag = conn.execute("SELECT id FROM tags WHERE name = ?", (name,)).fetchone()
        if tag:
            conn.execute("DELETE FROM tip_tags WHERE tip_id = ? AND tag_id = ?", (tip_id, tag["id"]))
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


def parse_batch(text):
    """Parse pasted text into tips.

    Each line is at least one tip (so plain, untagged tips paste one-per-line). Within a
    line, #tags attach to the current tip and a content word *after* a tag starts the next.

    e.g. "Drink water #health #morning Write goals #productivity" ->
         [("Drink water", ["health", "morning"]), ("Write goals", ["productivity"])]
         and a line with no #tags -> one tip with an empty tag list.
    """
    tips = []
    for line in text.splitlines():
        content_words = []
        tags = []
        for word in line.split():
            if word.startswith("#"):
                if len(word) > 1:
                    tags.append(word[1:].lower())
            else:
                if tags:  # a content word after tags starts a new tip
                    tips.append((" ".join(content_words).strip(), tags))
                    content_words, tags = [], []
                content_words.append(word)
        if content_words or tags:
            tips.append((" ".join(content_words).strip(), tags))
    return [(c, t) for c, t in tips if c]


@app.post("/api/tips/batch/preview")
@admin_required
def batch_preview():
    """Parse the pasted text into structured tips WITHOUT saving — for the review step."""
    data = request.get_json(force=True)
    text = data.get("text", "")
    return jsonify({"tips": [{"content": c, "tags": t} for c, t in parse_batch(text)]})


@app.post("/api/tips/batch/commit")
@admin_required
def batch_commit():
    """Insert the reviewed/edited list of tips. First tag is primary, rest secondary."""
    data = request.get_json(force=True)
    items = data.get("tips", [])
    inserted = []
    new_for_embed = []  # (tip_id, content, anecdote) for a single batched embed below
    skipped = 0

    with get_db() as conn:
        for item in items:
            content = (item.get("content") or "").strip()
            tags = [t.strip().lower() for t in item.get("tags", []) if t and t.strip()]
            if not content or not tags:
                skipped += 1  # need content and at least one (primary) tag
                continue
            cur = conn.execute("INSERT INTO tips (content) VALUES (?)", (content,))
            tip_id = cur.lastrowid
            for i, name in enumerate(tags):
                tag_id = get_or_create_tag(conn, name, tier="primary" if i == 0 else "secondary")
                conn.execute(
                    "INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)", (tip_id, tag_id)
                )
            inserted.append(tip_with_tags(conn, tip_id))
            new_for_embed.append((tip_id, content, ""))
        # Embed all the new tips in one batched call — best-effort, never blocks the import.
        if embeddings.is_enabled() and new_for_embed:
            try:
                embeddings.store_many(conn, new_for_embed)
            except llm.LLMError as e:
                app.logger.warning("batch embedding skipped (%d tips): %s", len(new_for_embed), e)
        conn.commit()

    return jsonify({"imported": len(inserted), "skipped": skipped}), 201


# ── Excel backup: export everything an admin has authored, and re-import it ──
# (content, anecdote, tags + their tier, video + times, and the "choose an angle" text).
# Votes / favorites are per-user activity, not authored content, so they're not included.
# One row per tip; the lens columns map to llm.ANALYSIS_LENSES.
EXPORT_BASE_COLUMNS = [
    "ID", "Content", "Anecdote", "Primary tags", "Secondary tags",
    "Video URL", "Video start (sec)", "Video end (sec)",
]
LENS_COLUMNS = [
    ("apply", "How to apply it"),
    ("avoid", "When not to apply it"),
    ("opposing", "Opposing wisdom"),
    ("misreadings", "Common misreadings"),
    ("figures", "Notable figures"),
]


@app.get("/api/tips/export")
@admin_required
def export_tips():
    """Stream every tip and its authored content as an .xlsx for offline backup."""
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Tips"
    ws.append(EXPORT_BASE_COLUMNS + [label for _, label in LENS_COLUMNS])

    with get_db() as conn:
        for tip in conn.execute("SELECT * FROM tips ORDER BY id").fetchall():
            tid = tip["id"]
            tagrows = conn.execute(
                "SELECT t.name, t.tier FROM tags t JOIN tip_tags tt ON t.id = tt.tag_id "
                "WHERE tt.tip_id = ? ORDER BY t.name", (tid,)
            ).fetchall()
            primary = ", ".join(r["name"] for r in tagrows if r["tier"] == "primary")
            secondary = ", ".join(r["name"] for r in tagrows if r["tier"] == "secondary")
            analysis = {
                r["lens"]: r["text"]
                for r in conn.execute(
                    "SELECT lens, text FROM tip_analysis WHERE tip_id = ?", (tid,)
                ).fetchall()
            }
            ws.append([
                tid, tip["content"], tip["anecdote"] or "", primary, secondary,
                tip["video_url"] or "", tip["video_start"] or 0, tip["video_end"] or 0,
            ] + [analysis.get(lens, "") for lens, _ in LENS_COLUMNS])

    # Roughly size the columns for readability.
    widths = [6, 60, 50, 22, 22, 36, 16, 16] + [40] * len(LENS_COLUMNS)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "practical-wisdom-tips-%s.xlsx" % time.strftime("%Y%m%d")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname,
    )


def _split_tags(value):
    """A cell of comma/semicolon-separated tag names → a clean lowercased list."""
    return [t.strip().lower() for t in re.split(r"[;,]", str(value or "")) if t.strip()]


def _cell_int(value):
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


@app.post("/api/tips/import")
@admin_required
def import_tips():
    """Re-import an exported .xlsx. Tips are matched by exact content: an existing tip is
    updated in place (so re-importing a backup is idempotent), a new one is created otherwise.
    Returns {created, updated, skipped}."""
    from openpyxl import load_workbook

    upload = request.files.get("file")
    if not upload:
        return jsonify({"error": "No file uploaded."}), 400
    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
    except Exception:
        return jsonify({"error": "Couldn't read that file — upload an .xlsx exported from here."}), 400
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    header = next(row_iter, None)
    if not header:
        return jsonify({"error": "The spreadsheet is empty."}), 400
    # Map header label → column index, so column order doesn't have to be exact.
    col_of = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

    def cell(row, *labels):
        for label in labels:
            i = col_of.get(label.lower())
            if i is not None and i < len(row) and row[i] is not None:
                return row[i]
        return None

    created = updated = skipped = 0
    for_embed = []
    with get_db() as conn:
        for row in row_iter:
            if row is None:
                continue
            content = str(cell(row, "Content") or "").strip()
            primary = _split_tags(cell(row, "Primary tags"))
            if not content or not primary:
                skipped += 1  # need content and at least one primary tag
                continue
            anecdote = str(cell(row, "Anecdote") or "").strip()
            video_url = str(cell(row, "Video URL") or "").strip()
            vstart = _cell_int(cell(row, "Video start (sec)", "Video start"))
            vend = _cell_int(cell(row, "Video end (sec)", "Video end"))
            secondary = _split_tags(cell(row, "Secondary tags"))

            existing = conn.execute("SELECT id FROM tips WHERE content = ?", (content,)).fetchone()
            if existing:
                tip_id = existing["id"]
                conn.execute(
                    "UPDATE tips SET anecdote = ?, video_url = ?, video_start = ?, video_end = ? WHERE id = ?",
                    (anecdote, video_url, vstart, vend, tip_id),
                )
                conn.execute("DELETE FROM tip_tags WHERE tip_id = ?", (tip_id,))  # rebuild from the sheet
                updated += 1
            else:
                cur = conn.execute(
                    "INSERT INTO tips (content, anecdote, video_url, video_start, video_end) VALUES (?, ?, ?, ?, ?)",
                    (content, anecdote, video_url, vstart, vend),
                )
                tip_id = cur.lastrowid
                created += 1

            for name in primary:
                conn.execute("INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)",
                             (tip_id, get_or_create_tag(conn, name, tier="primary")))
            for name in secondary:
                conn.execute("INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)",
                             (tip_id, get_or_create_tag(conn, name, tier="secondary")))
            for lens, label in LENS_COLUMNS:
                text = str(cell(row, label, lens) or "").strip()
                if text:
                    conn.execute(
                        "INSERT INTO tip_analysis (tip_id, lens, text) VALUES (?, ?, ?) "
                        "ON CONFLICT(tip_id, lens) DO UPDATE SET text = excluded.text",
                        (tip_id, lens, text),
                    )
                else:
                    conn.execute("DELETE FROM tip_analysis WHERE tip_id = ? AND lens = ?", (tip_id, lens))
            for_embed.append((tip_id, content, anecdote))

        if embeddings.is_enabled() and for_embed:
            try:
                embeddings.store_many(conn, for_embed)
            except Exception as e:  # best-effort: never let embedding break a restore
                app.logger.warning("import embedding skipped (%d tips): %s", len(for_embed), e)
        conn.commit()

    return jsonify({"created": created, "updated": updated, "skipped": skipped})


@app.post("/api/llm/suggest-tags")
@admin_required
def llm_suggest_tags():
    """Suggest tags (from the existing taxonomy) for a list of tip contents, via Gemini."""
    if not llm.is_enabled():
        return jsonify({"error": "AI tagging isn't configured. Set GEMINI_API_KEY in .env."}), 503
    data = request.get_json(force=True) or {}
    contents = [c for c in (data.get("contents") or []) if (c or "").strip()]
    if not contents:
        return jsonify({"suggestions": []})
    with get_db() as conn:
        primary = [r["name"] for r in conn.execute("SELECT name FROM tags WHERE tier = 'primary' ORDER BY name")]
        secondary = [r["name"] for r in conn.execute("SELECT name FROM tags WHERE tier = 'secondary' ORDER BY name")]
    if not primary:
        return jsonify({"error": "Add some primary tags first so the AI has a taxonomy to use."}), 400
    try:
        suggestions = llm.suggest_tags_batch(contents, primary, secondary)
    except llm.LLMError as e:
        return jsonify({"error": "AI tagging failed: %s" % e}), 502
    return jsonify({"suggestions": suggestions})


@app.get("/api/embeddings/status")
@admin_required
def embeddings_status():
    """How many tips currently have a semantic embedding (drives the 'rebuild' button)."""
    with get_db() as conn:
        return jsonify(embeddings.status(conn))


@app.post("/api/embeddings/rebuild")
@admin_required
def embeddings_rebuild():
    """Embed every tip that's missing or out of date. Safe to run repeatedly."""
    if not embeddings.is_enabled():
        return jsonify({"error": "Embeddings need an API key. Set GEMINI_API_KEY in .env."}), 503
    with get_db() as conn:
        try:
            result = embeddings.sync_all(conn)
        except llm.LLMError as e:
            return jsonify({"error": "Embedding failed: %s" % e}), 502
    return jsonify(result)


@app.get("/api/tags")
def list_tags():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT t.name, t.tier, COUNT(tt.tip_id) as count FROM tags t LEFT JOIN tip_tags tt ON t.id = tt.tag_id GROUP BY t.id ORDER BY t.name"
        ).fetchall()
        return jsonify(
            [{"name": r["name"], "tier": r["tier"], "count": r["count"]} for r in rows]
        )


@app.delete("/api/tags/<name>")
@admin_required
def delete_tag(name):
    name = name.strip().lower()
    with get_db() as conn:
        row = conn.execute("SELECT id FROM tags WHERE name = ?", (name,)).fetchone()
        if not row:
            return jsonify({"error": "tag not found"}), 404
        affected = conn.execute(
            "SELECT COUNT(*) AS n FROM tip_tags WHERE tag_id = ?", (row["id"],)
        ).fetchone()["n"]
        conn.execute("DELETE FROM tags WHERE id = ?", (row["id"],))
        conn.commit()
        return jsonify({"deleted": name, "tips_affected": affected})


@app.post("/api/tags/batch")
@admin_required
def import_tags():
    """Add tags to the allowed list from comma-separated text (no tip attached).

    Accepts an optional "tier" ('primary' or 'secondary', default 'primary').
    Existing tags are reclassified to the chosen tier.
    """
    data = request.get_json(force=True)
    text = data.get("text", "")
    tier = data.get("tier", "primary")
    if tier not in ("primary", "secondary"):
        tier = "primary"
    names = [t.replace("#", "").strip().lower() for t in text.split(",")]
    names = [n for n in names if n]

    added = 0
    with get_db() as conn:
        for name in names:
            existed = conn.execute("SELECT 1 FROM tags WHERE name = ?", (name,)).fetchone()
            get_or_create_tag(conn, name, tier=tier)
            if not existed:
                added += 1
        conn.commit()

    return jsonify({"added": added, "submitted": len(names), "tier": tier}), 201


# ──────────────────────── Auth & per-user actions ────────────────────────

@app.get("/api/me")
def api_me():
    """Who is logged in (or null), plus whether Google login is configured at all."""
    uid = current_user_id()
    user = None
    if uid:
        with get_db() as conn:
            u = conn.execute(
                "SELECT id, email, name, picture FROM users WHERE id = ?", (uid,)
            ).fetchone()
        if u:
            user = {"id": u["id"], "email": u["email"], "name": u["name"], "picture": u["picture"]}
        else:
            session.pop("uid", None)  # stale session (user row gone)
    pending_submissions = 0
    if is_admin():
        with get_db() as conn:
            pending_submissions = conn.execute(
                "SELECT COUNT(*) AS n FROM tip_submissions WHERE status = 'pending'").fetchone()["n"]
    return jsonify({"user": user, "auth_enabled": AUTH_ENABLED, "is_admin": is_admin(),
                    "llm_enabled": llm.is_enabled(), "embeddings_enabled": embeddings.is_enabled(),
                    "pending_submissions": pending_submissions, "csrf_token": csrf_token()})


@app.post("/api/admin/login")
def admin_login():
    ip = request.remote_addr or "?"
    if _login_blocked(ip):
        return jsonify({"error": "Too many attempts — wait a few minutes and try again."}), 429
    data = request.get_json(force=True) or {}
    ok = (secrets.compare_digest(data.get("username", ""), ADMIN_USERNAME)
          and check_password_hash(ADMIN_PASSWORD_HASH, data.get("password", "")))
    if not ok:
        _login_failed(ip)
        return jsonify({"error": "Invalid administrator credentials."}), 401
    _login_attempts.pop(ip, None)  # reset on success
    session["is_admin"] = True
    return jsonify({"is_admin": True})


@app.post("/api/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return jsonify({"is_admin": False})


@app.get("/login")
def login():
    if not AUTH_ENABLED:
        return "Google login is not configured (set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET).", 503
    return oauth.google.authorize_redirect(url_for("auth_callback", _external=True))


@app.get("/auth/callback")
def auth_callback():
    if not AUTH_ENABLED:
        return redirect("/")
    token = oauth.google.authorize_access_token()  # exchanges code, verifies the ID token
    info = token.get("userinfo") or {}
    sub = info.get("sub")
    if not sub:
        return "Sign-in failed.", 400
    with get_db() as conn:
        row = conn.execute("SELECT id FROM users WHERE google_sub = ?", (sub,)).fetchone()
        if row:
            uid = row["id"]
            conn.execute(
                "UPDATE users SET email = ?, name = ?, picture = ? WHERE id = ?",
                (info.get("email"), info.get("name"), info.get("picture"), uid),
            )
        else:
            cur = conn.execute(
                "INSERT INTO users (google_sub, email, name, picture) VALUES (?, ?, ?, ?)",
                (sub, info.get("email"), info.get("name"), info.get("picture")),
            )
            uid = cur.lastrowid
        conn.commit()
    session["uid"] = uid
    return redirect("/")


@app.post("/logout")
def logout():
    session.pop("uid", None)
    return jsonify({"ok": True})


@app.post("/api/tips/<int:tip_id>/vote")
def vote_tip(tip_id):
    """Set the current user's vote to +1, -1, or 0 (0 removes it). Returns the updated tip."""
    uid = current_user_id()
    if not uid:
        return jsonify({"error": "Sign in to vote."}), 401
    value = (request.get_json(force=True) or {}).get("value", 0)
    if value not in (1, -1, 0):
        return jsonify({"error": "value must be 1, -1, or 0"}), 400
    with get_db() as conn:
        if not conn.execute("SELECT 1 FROM tips WHERE id = ?", (tip_id,)).fetchone():
            return jsonify({"error": "tip not found"}), 404
        if value == 0:
            conn.execute("DELETE FROM votes WHERE user_id = ? AND tip_id = ?", (uid, tip_id))
        else:
            conn.execute(
                """INSERT INTO votes (user_id, tip_id, value) VALUES (?, ?, ?)
                   ON CONFLICT(user_id, tip_id) DO UPDATE SET value = excluded.value""",
                (uid, tip_id, value),
            )
        conn.commit()
        return jsonify(tip_with_tags(conn, tip_id))


@app.post("/api/favorites/insights")
def favorites_insights():
    """Reflect on the signed-in user's favourite tips: themes, what resonates, next steps.

    Favourites = the user's upvoted tips. Needs a signed-in user and at least 3 favourites
    (so there's a pattern to read). Returns {"count": n, "insight": {...}}.
    """
    uid = current_user_id()
    if not uid:
        return jsonify({"error": "Sign in to reflect on your favourites."}), 401
    if not llm.is_enabled():
        return jsonify({"error": "This needs an AI key (set GEMINI_API_KEY)."}), 503
    with get_db() as conn:
        rows = conn.execute(
            """SELECT t.id FROM tips t JOIN votes v ON v.tip_id = t.id
               WHERE v.user_id = ? AND v.value = 1
               ORDER BY t.id DESC LIMIT 80""", (uid,)).fetchall()
        tips = [tip_with_tags(conn, r["id"]) for r in rows]
        library_size = conn.execute("SELECT COUNT(*) AS n FROM tips").fetchone()["n"]
    if len(tips) < 3:
        return jsonify({"error": "Save at least 3 favourites first, so there's a pattern to read."}), 400
    try:
        insight = llm.reflect_on_favorites(
            [{"content": t["content"], "tags": t["tags"]} for t in tips], library_size=library_size)
    except llm.LLMError as e:
        return jsonify({"error": "Reflection failed: %s" % e}), 502
    return jsonify({"count": len(tips), "insight": insight})


# ──────────────────────── Community tip submissions ────────────────────────

def _submission_json(conn, row):
    submitter = None
    if row["user_id"]:
        u = conn.execute("SELECT name, email FROM users WHERE id = ?", (row["user_id"],)).fetchone()
        if u:
            submitter = u["name"] or u["email"]
    return {
        "id": row["id"], "content": row["content"], "anecdote": row["anecdote"] or "",
        "tags": [t for t in (row["tags"] or "").split(",") if t],
        "status": row["status"], "created_at": row["created_at"], "submitter": submitter,
    }


@app.post("/api/submissions")
def create_submission():
    """A signed-in user suggests a tip. It enters the moderation queue (status=pending)."""
    uid = current_user_id()
    if not uid:
        return jsonify({"error": "Sign in to suggest a tip."}), 401
    data = request.get_json(force=True) or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "Write the tip first."}), 400
    if len(content) > 600:
        return jsonify({"error": "Keep tips under 600 characters."}), 400
    anecdote = (data.get("anecdote") or "").strip()[:1000]
    tags = ",".join(t.strip().lower() for t in (data.get("tags") or []) if t and t.strip())
    with get_db() as conn:
        pending = conn.execute(
            "SELECT COUNT(*) AS n FROM tip_submissions WHERE user_id = ? AND status = 'pending'",
            (uid,)).fetchone()["n"]
        if pending >= 25:
            return jsonify({"error": "You have several tips awaiting review — give those a chance first."}), 429
        cur = conn.execute(
            "INSERT INTO tip_submissions (content, anecdote, tags, user_id) VALUES (?, ?, ?, ?)",
            (content, anecdote, tags, uid))
        conn.commit()
        row = conn.execute("SELECT * FROM tip_submissions WHERE id = ?", (cur.lastrowid,)).fetchone()
        return jsonify(_submission_json(conn, row)), 201


@app.get("/api/submissions/mine")
def my_submissions():
    """The signed-in user's own submissions, with their review status (pending/approved/rejected)."""
    uid = current_user_id()
    if not uid:
        return jsonify({"submissions": []})
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM tip_submissions WHERE user_id = ? ORDER BY created_at DESC LIMIT 50", (uid,)).fetchall()
        return jsonify({"submissions": [_submission_json(conn, r) for r in rows]})


@app.get("/api/submissions")
@admin_required
def list_submissions():
    """Admin: list submissions (default the pending queue)."""
    status = request.args.get("status", "pending")
    with get_db() as conn:
        if status == "all":
            rows = conn.execute("SELECT * FROM tip_submissions ORDER BY created_at DESC").fetchall()
        else:
            if status not in ("pending", "approved", "rejected"):
                status = "pending"
            rows = conn.execute(
                "SELECT * FROM tip_submissions WHERE status = ? ORDER BY created_at DESC", (status,)).fetchall()
        return jsonify({"submissions": [_submission_json(conn, r) for r in rows]})


@app.post("/api/submissions/<int:sub_id>/approve")
@admin_required
def approve_submission(sub_id):
    """Approve a submission: create a real tip (with tags + embedding) and mark it approved.

    The admin may pass edited content/anecdote/tags; otherwise the submitted values are used.
    First tag becomes primary, the rest secondary (same as a batch import).
    """
    data = request.get_json(force=True) or {}
    with get_db() as conn:
        sub = conn.execute("SELECT * FROM tip_submissions WHERE id = ?", (sub_id,)).fetchone()
        if not sub:
            return jsonify({"error": "submission not found"}), 404
        if sub["status"] != "pending":
            return jsonify({"error": "Already %s." % sub["status"]}), 409
        content = (data.get("content") or sub["content"] or "").strip()
        anecdote = (data.get("anecdote") if data.get("anecdote") is not None else (sub["anecdote"] or "")).strip()
        if "tags" in data:
            tags = [t.strip().lower() for t in data.get("tags") if t and t.strip()]
        else:
            tags = [t for t in (sub["tags"] or "").split(",") if t]
        if not content:
            return jsonify({"error": "content is required"}), 400
        if not tags:
            return jsonify({"error": "Add at least one tag (the first becomes the primary)."}), 400
        cur = conn.execute("INSERT INTO tips (content, anecdote) VALUES (?, ?)", (content, anecdote))
        tip_id = cur.lastrowid
        for i, name in enumerate(tags):
            tag_id = get_or_create_tag(conn, name, tier="primary" if i == 0 else "secondary")
            conn.execute("INSERT OR IGNORE INTO tip_tags (tip_id, tag_id) VALUES (?, ?)", (tip_id, tag_id))
        embed_quietly(conn, tip_id, content, anecdote)   # index it for semantic search
        conn.execute(
            "UPDATE tip_submissions SET status='approved', reviewed_at=CURRENT_TIMESTAMP, tip_id=? WHERE id=?",
            (tip_id, sub_id))
        conn.commit()
        return jsonify({"approved": sub_id, "tip": tip_with_tags(conn, tip_id)}), 201


@app.post("/api/submissions/<int:sub_id>/reject")
@admin_required
def reject_submission(sub_id):
    """Admin: reject a submission (kept in the table as a record, not turned into a tip)."""
    with get_db() as conn:
        if not conn.execute("SELECT 1 FROM tip_submissions WHERE id = ?", (sub_id,)).fetchone():
            return jsonify({"error": "submission not found"}), 404
        conn.execute(
            "UPDATE tip_submissions SET status='rejected', reviewed_at=CURRENT_TIMESTAMP WHERE id = ?", (sub_id,))
        conn.commit()
    return jsonify({"rejected": sub_id})


@app.get("/api/seen")
def get_seen():
    """Tip ids the current user has already visited (empty when not signed in)."""
    uid = current_user_id()
    if not uid:
        return jsonify({"seen": []})
    with get_db() as conn:
        rows = conn.execute("SELECT tip_id FROM seen_tips WHERE user_id = ?", (uid,)).fetchall()
    return jsonify({"seen": [r["tip_id"] for r in rows]})


@app.post("/api/tips/<int:tip_id>/seen")
def mark_seen(tip_id):
    """Record that the current user has visited this tip. No-op when not signed in."""
    uid = current_user_id()
    if not uid:
        return jsonify({"ok": False})
    with get_db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO seen_tips (user_id, tip_id) VALUES (?, ?)", (uid, tip_id)
        )
        conn.commit()
    return jsonify({"ok": True})


@app.post("/api/seen/reset")
def reset_seen():
    """Clear the current user's visited history (so they can re-explore from scratch)."""
    uid = current_user_id()
    if uid:
        with get_db() as conn:
            conn.execute("DELETE FROM seen_tips WHERE user_id = ?", (uid,))
            conn.commit()
    return jsonify({"ok": True})


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=int(os.environ.get("PORT", "5001")))

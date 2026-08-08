"""Generate one illustration per PRODUCTION tip, all in the chosen library style.

Costs credits (5 per image on Gemini 2.5 Flash). Source of truth is a production
Excel export, because the live database is only reachable through the deployed app —
images are named by the PRODUCTION tip id so they line up after deploy.

    python3 scripts/gen_tip_images.py --export path/to/export.xlsx [--style goldline]
    python3 scripts/gen_tip_images.py --export ... --limit 3      # small trial first
    python3 scripts/gen_tip_images.py --export ... --dry-run      # prompts only, no spend

Resumable: a tip whose image already exists on disk is skipped, so an interrupted run
costs nothing to continue and a later top-up only pays for genuinely new tips.
"""
import argparse
import json
import os
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv()

import imagegen
import llm

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "static", "tip_images")
CONCEPTS = os.path.join(ROOT, "scripts", "tip_concepts.json")   # cached, so a re-run is free


def load_export(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    head = [str(x).strip() if x else "" for x in rows[0]]
    i_id, i_c = head.index("ID"), head.index("Content")
    out = []
    for r in rows[1:]:
        if len(r) > i_c and r[i_c] and str(r[i_c]).strip():
            out.append((int(r[i_id]), str(r[i_c]).strip()))
    return out


def concept_for(content):
    """A one-line VISUAL metaphor for a tip — never literal text in the picture.

    Uses the LLM already wired into the app. Falls back to the tip's own words, which
    still produces a usable image, just a more literal one.
    """
    prompt = (
        "Turn this piece of practical wisdom into ONE short visual metaphor that an "
        "illustrator could draw. Reply with a single line: concrete objects and a simple "
        "scene, 12-20 words, no people's faces, no written words or lettering in the "
        "scene, no abstract nouns.\n\n"
        'Wisdom: "%s"\n\n'
        'Return JSON: {"concept": "..."}'
    ) % content
    try:
        parsed = llm._complete_json(prompt, temperature=0.7)
        c = (parsed or {}).get("concept", "").strip()
        if c:
            return c
    except Exception:
        pass
    return content


def load_concept_cache():
    if os.path.exists(CONCEPTS):
        with open(CONCEPTS) as fh:
            return json.load(fh)
    return {}


def save_concept_cache(cache):
    with open(CONCEPTS, "w") as fh:
        json.dump(cache, fh, indent=1, sort_keys=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--export", required=True, help="production .xlsx export")
    ap.add_argument("--style", default=os.environ.get("TIP_IMAGE_STYLE", "goldline"))
    ap.add_argument("--limit", type=int, default=0, help="only do the first N (trial run)")
    ap.add_argument("--dry-run", action="store_true", help="print prompts, spend nothing")
    ap.add_argument("--fresh-concepts", action="store_true",
                    help="re-derive concepts for tips still missing an image (use when a "
                         "retry keeps failing — a different metaphor usually gets through)")
    args = ap.parse_args()

    if args.style not in imagegen.STYLE_TEMPLATES:
        sys.exit("unknown style: %s (have %s)" % (args.style, list(imagegen.STYLE_TEMPLATES)))
    if not args.dry_run and not imagegen.is_enabled():
        sys.exit("THREEDAI_API_KEY is not set")

    tips = load_export(args.export)
    os.makedirs(OUT, exist_ok=True)
    todo = [(i, c) for i, c in tips if not os.path.exists(os.path.join(OUT, "%d.webp" % i))]
    if args.limit:
        todo = todo[:args.limit]

    done_count = len(tips) - len([t for t in tips
                                  if not os.path.exists(os.path.join(OUT, "%d.webp" % t[0]))])
    # Cost differs wildly by backend, so quote the one actually in use rather than a
    # hardcoded number (the old "5 credits" figure was wrong by ~6x and cost real money).
    if imagegen.provider() == "gemini":
        est = "~$%.2f at Gemini's ~$0.04/image" % (len(todo) * 0.04)
    else:
        est = "~%d credits at 3D AI Studio's observed ~30/image" % (len(todo) * 30)
    print("style=%s | backend=%s | tips in export=%d | already done=%d | to generate=%d | %s"
          % (args.style, imagegen.provider(), len(tips), done_count, len(todo), est), flush=True)
    if not todo:
        print("nothing to do"); return

    cache = load_concept_cache()
    if args.fresh_concepts:
        # These tips have no image, so their cached metaphor is the prime suspect —
        # drop it and let a new one be derived.
        for tip_id, _ in todo:
            cache.pop(str(tip_id), None)
        save_concept_cache(cache)
        print("cleared cached concepts for %d unfinished tips" % len(todo))
    ok = failed = 0
    started = time.time()

    def note_wait(secs, attempt):
        print("      throttled, waiting %ss" % secs, flush=True)

    for n, (tip_id, content) in enumerate(todo, 1):
        key = str(tip_id)
        if key not in cache:
            cache[key] = concept_for(content)
            save_concept_cache(cache)          # cache as we go; concepts are reusable
        prompt = imagegen.build_prompt(args.style, cache[key])

        if args.dry_run:
            print("[%3d/%d] tip %s\n        concept: %s" % (n, len(todo), tip_id, cache[key]))
            continue

        mins = (time.time() - started) / 60
        print("[%3d/%d] tip %-4s (%.0f min elapsed) %s" % (n, len(todo), tip_id, mins,
                                                           content[:44]), flush=True)
        try:
            blob = imagegen.generate(prompt, aspect_ratio="4:3", resolution="1K",
                                     on_wait=note_wait)
            with open(os.path.join(OUT, "%d.webp" % tip_id), "wb") as fh:
                fh.write(blob)
            ok += 1
            print("         saved %5.0f KB" % (len(blob) / 1024), flush=True)
        except imagegen.ImageGenError as e:
            failed += 1
            print("         FAILED: %s" % e, flush=True)

    print("\ndone: %d generated, %d failed, %.0f min total" % (ok, failed,
                                                              (time.time() - started) / 60))
    if failed:
        print("re-run the same command to retry only the failures (finished ones are skipped)")


if __name__ == "__main__":
    main()
